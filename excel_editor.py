import customtkinter as ctk
import pandas as pd
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox
from tksheet import Sheet

import database


class ExcelEditor(ctk.CTkFrame):
    def __init__(self, parent, current_user):
        super().__init__(parent)
        self.parent = parent
        self.current_user = current_user
        self.df = None
        self._resize_after_id = None

        self.top_bar = ctk.CTkFrame(self, height=50)
        self.top_bar.pack(side="top", fill="x", padx=10, pady=(10, 5))

        self.save_button = ctk.CTkButton(
            self.top_bar,
            text="Save Changes",
            command=self.save_file,
            state="disabled",
        )
        self.save_button.pack(side="left", padx=5)

        self.refresh_button = ctk.CTkButton(
            self.top_bar,
            text="Refresh",
            command=self.load_table,
            fg_color="#6c757d",
            hover_color="#5a6268",
            width=90,
        )
        self.refresh_button.pack(side="left", padx=5)

        self.import_button = ctk.CTkButton(
            self.top_bar,
            text="Import Excel",
            command=self.import_excel,
            fg_color="#0d6efd",
            hover_color="#0b5ed7",
            width=110,
        )
        self.import_button.pack(side="left", padx=5)

        self.export_button = ctk.CTkButton(
            self.top_bar,
            text="Export Excel",
            command=self.export_excel,
            fg_color="#28a745",
            hover_color="#218838",
            width=110,
        )
        self.export_button.pack(side="left", padx=5)

        self.input_frame = ctk.CTkFrame(self)
        self.input_frame.pack(side="top", fill="x", padx=10, pady=(0, 5))

        self.entries = {}
        input_cols = ["Rep", "Date", "Reçu", "Libelles", "Versement", "Débours"]

        for col in input_cols:
            container = ctk.CTkFrame(self.input_frame, fg_color="transparent")
            container.pack(side="left", padx=4, pady=4, expand=True, fill="x")
            ctk.CTkLabel(container, text=col, font=ctk.CTkFont(size=11)).pack(anchor="w")
            entry = ctk.CTkEntry(container, height=26)
            entry.pack(fill="x")
            self.entries[col] = entry

        self.add_btn = ctk.CTkButton(
            self.input_frame,
            text="Add",
            command=self.add_from_inputs,
            fg_color="#17a2b8",
            hover_color="#138496",
            width=70,
        )
        self.add_btn.pack(side="left", padx=10, pady=(18, 0))

        self.sheet_frame = ctk.CTkFrame(self)
        self.sheet_frame.pack(side="top", fill="both", expand=True, padx=10, pady=(0, 5))

        self.sheet = Sheet(
            self.sheet_frame,
            show_x_scrollbar=True,
            show_y_scrollbar=True,
        )
        self.sheet.enable_bindings(
            (
                "single_select",
                "drag_select",
                "column_select",
                "row_select",
                "column_width_resize",
                "double_click_column_resize",
                "row_height_resize",
                "double_click_row_resize",
                "right_click_popup_menu",
                "rc_select",
                "rc_insert_row",
                "rc_delete_row",
                "copy",
                "cut",
                "paste",
                "delete",
                "undo",
                "edit_cell",
            )
        )
        self.sheet.extra_bindings("end_edit_cell", func=self.on_cell_edited)
        self.sheet.pack(fill="both", expand=True)

        self.sheet_frame.bind("<Configure>", self._on_frame_configure)

        self.footer = ctk.CTkFrame(self, height=50)
        self.footer.pack(side="bottom", fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(
            self.footer,
            text="TOTAUX",
            font=ctk.CTkFont(size=13, weight="bold"),
        ).pack(side="left", padx=16)

        self.total_labels = {}
        for col in ("Versement", "Débours", "Honoraires", "T.V.A", "Honors/H.T"):
            container = ctk.CTkFrame(self.footer, fg_color="transparent")
            container.pack(side="left", padx=14, expand=True)
            ctk.CTkLabel(container, text=col, font=ctk.CTkFont(size=10, weight="bold")).pack()
            label = ctk.CTkLabel(
                container,
                text="0,00",
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color="#28a745",
            )
            label.pack()
            self.total_labels[col] = label

        self.load_table()

    def load_table(self):
        try:
            headers, data = database.get_table_data()
            self.df = pd.DataFrame(data, columns=headers)

            self.sheet.set_sheet_data(data)
            self.sheet.headers(headers)

            self.save_button.configure(state="normal")

            self.after(100, self._fit_columns)
            self.recalculate()
        except Exception as exc:
            messagebox.showerror("Erreur", f"Impossible de charger la table:\n{exc}")

    def _on_frame_configure(self, event):
        if self._resize_after_id:
            self.after_cancel(self._resize_after_id)
        self._resize_after_id = self.after(80, lambda: self._fit_columns(event.width))

    def _fit_columns(self, frame_width=None):
        headers = self.sheet.headers()
        if not headers:
            return
        column_count = len(headers)
        if frame_width is None:
            self.sheet_frame.update_idletasks()
            frame_width = self.sheet_frame.winfo_width()

        available = max(frame_width - 68, column_count * 60)
        column_width = available // column_count

        try:
            self.sheet.set_column_widths([column_width] * column_count)
        except Exception:
            pass

    def add_from_inputs(self):
        headers = self.sheet.headers()
        if not headers:
            return

        new_row = [""] * len(headers)

        if "N°Fact" in headers:
            fact_index = headers.index("N°Fact")
            data = self.sheet.get_sheet_data()
            last_fact = 0
            for row in data:
                value = str(row[fact_index]).strip()
                try:
                    last_fact = max(last_fact, int(value))
                except ValueError:
                    pass
            new_row[fact_index] = f"{last_fact + 1:02d}"

        for col_name, entry in self.entries.items():
            if col_name in headers:
                new_row[headers.index(col_name)] = entry.get().strip()
                entry.delete(0, "end")

        data = self.sheet.get_sheet_data()
        data.append(new_row)
        self.sheet.set_sheet_data(data)
        self.recalculate()

    def on_cell_edited(self, event=None):
        self.recalculate()

    def recalculate(self, event=None):
        headers = self.sheet.headers()
        if not headers:
            return

        try:
            versement_index = headers.index("Versement")
            debours_index = headers.index("Débours")
            honoraires_index = headers.index("Honoraires")
            tva_index = headers.index("T.V.A")
            ht_index = headers.index("Honors/H.T")
        except ValueError:
            return

        data = self.sheet.get_sheet_data()

        for row_index, row in enumerate(data):
            try:
                versement_raw = str(row[versement_index]).replace(",", ".").replace(" ", "").strip()
                debours_raw = str(row[debours_index]).replace(",", ".").replace(" ", "").strip()
                if not versement_raw and not debours_raw:
                    continue

                versement = float(versement_raw) if versement_raw else 0.0
                debours = float(debours_raw) if debours_raw else 0.0
                honoraires = versement - debours
                ht = honoraires / 1.2
                tva = honoraires - ht

                for column_index, value in (
                    (honoraires_index, honoraires),
                    (tva_index, tva),
                    (ht_index, ht),
                ):
                    formatted = self._format_money(value)
                    if str(row[column_index]) != formatted:
                        self.sheet.set_cell_data(row_index, column_index, formatted)
            except (ValueError, IndexError):
                pass

        totals = {column: 0.0 for column in self.total_labels}
        for row in self.sheet.get_sheet_data():
            for column_name in totals:
                try:
                    column_index = headers.index(column_name)
                    value = str(row[column_index]).replace(",", ".").replace(" ", "").strip()
                    if value:
                        totals[column_name] += float(value)
                except (ValueError, IndexError):
                    pass

        for column_name, value in totals.items():
            self.total_labels[column_name].configure(text=self._format_money(value))

    def _format_money(self, value):
        return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")

    def save_file(self):
        try:
            data = self.sheet.get_sheet_data()
            headers = self.sheet.headers()
            new_df = pd.DataFrame(data, columns=headers)

            if self.df is not None:
                self._log_changes(self.df, new_df, headers)

            database.save_table_data(headers, data)
            self.df = new_df
            messagebox.showinfo("Succès", "Table sauvegardée dans la base de données!")
        except Exception as exc:
            messagebox.showerror("Erreur", f"Impossible de sauvegarder:\n{exc}")

    def import_excel(self):
        current_headers = self.sheet.headers()
        if not current_headers:
            messagebox.showwarning("Erreur", "Aucune table chargée.")
            return

        filepath = filedialog.askopenfilename(
            title="Importer un fichier",
            filetypes=[
                ("Supported files", "*.xlsx *.xls *.csv *.tsv *.txt"),
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("Text files", "*.tsv *.txt"),
                ("All files", "*.*"),
            ],
        )
        if not filepath:
            return

        try:
            imported_df = self._read_import_dataframe(filepath, current_headers)
            imported_headers = imported_df.columns.tolist()

            if not self._same_table_headers(current_headers, imported_headers):
                messagebox.showerror(
                    "Table incorrecte",
                    "Le fichier importé n'a pas la même structure.\n\n"
                    f"Attendu:\n{self._headers_for_message(current_headers)}\n\n"
                    f"Reçu:\n{self._headers_for_message(imported_headers)}",
                )
                return

            imported_df = self._clean_import_dataframe(imported_df, current_headers)
            data = imported_df.values.tolist()
            self.sheet.set_sheet_data(data)
            self.sheet.headers(current_headers)
            self.recalculate()
            self.after(100, self._fit_columns)
            messagebox.showinfo(
                "Import Excel",
                "Fichier importé avec succès.\nVérifie ou modifie les données, puis clique sur Save Changes.",
            )
        except Exception as exc:
            messagebox.showerror("Erreur", f"Impossible d'importer:\n{exc}")

    def export_excel(self):
        headers = self.sheet.headers()
        if not headers:
            messagebox.showwarning("Erreur", "Aucune table à exporter.")
            return

        filepath = filedialog.asksaveasfilename(
            title="Exporter vers Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="export_table.xlsx",
        )
        if not filepath:
            return

        try:
            data = self.sheet.get_sheet_data()
            export_df = self._prepare_export_dataframe(data, headers)
            self._write_styled_excel(filepath, export_df)
            messagebox.showinfo("Succès", f"Export Excel créé:\n{filepath}")
        except Exception as exc:
            messagebox.showerror("Erreur", f"Impossible d'exporter:\n{exc}")

    def _same_table_headers(self, expected_headers, imported_headers):
        if len(expected_headers) != len(imported_headers):
            return False
        expected = [self._normal_header(header) for header in expected_headers]
        imported = [self._normal_header(header) for header in imported_headers]
        return expected == imported

    def _headers_for_message(self, headers):
        return " | ".join(str(header) for header in headers)

    def _read_import_dataframe(self, filepath, current_headers):
        direct_df = self._read_tabular_file(filepath).fillna("")
        if self._same_table_headers(current_headers, direct_df.columns.tolist()):
            return direct_df

        raw_df = self._read_tabular_file(filepath, header=None).fillna("")
        expected = [self._normal_header(header) for header in current_headers]
        header_count = len(current_headers)
        scan_rows = min(len(raw_df), 25)

        for row_index in range(scan_rows):
            row_values = raw_df.iloc[row_index].tolist()[:header_count]
            normalized = [self._normal_header(value) for value in row_values]
            if normalized == expected:
                imported_df = self._read_tabular_file(filepath, header=row_index).fillna("")
                imported_df = imported_df.iloc[:, :header_count]
                imported_df.columns = current_headers
                return imported_df

        return direct_df

    def _read_tabular_file(self, filepath, header=0):
        extension = Path(filepath).suffix.lower()
        if extension in {".xlsx", ".xls", ".xlsm"}:
            return pd.read_excel(filepath, header=header, dtype=str)
        if extension == ".tsv":
            return pd.read_csv(filepath, header=header, dtype=str, sep="\t")
        if extension in {".csv", ".txt"}:
            return pd.read_csv(filepath, header=header, dtype=str, sep=None, engine="python")

        try:
            return pd.read_excel(filepath, header=header, dtype=str)
        except Exception:
            return pd.read_csv(filepath, header=header, dtype=str, sep=None, engine="python")

    def _clean_import_dataframe(self, imported_df, current_headers):
        cleaned_df = imported_df.iloc[:, : len(current_headers)].copy().fillna("")
        cleaned_df.columns = current_headers

        first_column = current_headers[0]
        total_mask = cleaned_df[first_column].apply(lambda value: self._normal_header(value) == "totaux")
        return cleaned_df.loc[~total_mask]

    def _prepare_export_dataframe(self, data, headers):
        export_df = pd.DataFrame(data, columns=headers)
        for column in self._money_columns(headers):
            export_df[column] = export_df[column].apply(self._parse_money)
        return export_df

    def _write_styled_excel(self, filepath, export_df):
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        sheet_name = "Tableau"
        header_row = 4
        data_start_row = header_row + 1
        row_count = len(export_df)
        total_row = data_start_row + row_count

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            export_df.to_excel(writer, sheet_name=sheet_name, startrow=header_row - 1, index=False)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            max_col = max(1, len(export_df.columns))
            max_col_letter = get_column_letter(max_col)

            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = "Export Tableau"
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill("solid", fgColor="1F4E78")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            worksheet.row_dimensions[1].height = 30

            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
            subtitle_cell = worksheet.cell(row=2, column=1)
            subtitle_cell.value = f"Exporté le {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            subtitle_cell.font = Font(size=10, italic=True, color="666666")
            subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

            header_fill = PatternFill("solid", fgColor="245C8A")
            band_fill = PatternFill("solid", fgColor="F4F8FB")
            total_fill = PatternFill("solid", fgColor="D9EAF7")
            thin_side = Side(style="thin", color="B7C9D6")
            medium_side = Side(style="medium", color="1F4E78")
            border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            header_border = Border(left=thin_side, right=thin_side, top=medium_side, bottom=medium_side)

            for cell in worksheet[header_row]:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = header_border
            worksheet.row_dimensions[header_row].height = 24

            for row_idx in range(data_start_row, total_row):
                if (row_idx - data_start_row) % 2 == 1:
                    for cell in worksheet[row_idx]:
                        cell.fill = band_fill
                for cell in worksheet[row_idx]:
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

            money_columns = set(self._money_columns(export_df.columns))
            for col_idx, header in enumerate(export_df.columns, start=1):
                letter = get_column_letter(col_idx)
                if header in money_columns:
                    for row_idx in range(data_start_row, total_row):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                elif self._normal_header(header) in {"n_fact", "rep", "date", "recu"}:
                    for row_idx in range(data_start_row, total_row):
                        worksheet.cell(row=row_idx, column=col_idx).alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                        )
                worksheet.column_dimensions[letter].width = self._export_column_width(header, export_df[header])

            worksheet.cell(row=total_row, column=1).value = "TOTAUX"
            for col_idx, header in enumerate(export_df.columns, start=1):
                cell = worksheet.cell(row=total_row, column=col_idx)
                cell.fill = total_fill
                cell.font = Font(bold=True, color="1F4E78")
                cell.border = header_border
                cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "right", vertical="center")
                if header in money_columns:
                    letter = get_column_letter(col_idx)
                    cell.value = f"=SUM({letter}{data_start_row}:{letter}{total_row - 1})" if row_count else 0
                    cell.number_format = '#,##0.00'
            worksheet.row_dimensions[total_row].height = 24

            worksheet.auto_filter.ref = (
                f"A{header_row}:{max_col_letter}{total_row - 1}"
                if row_count
                else f"A{header_row}:{max_col_letter}{header_row}"
            )
            worksheet.freeze_panes = f"A{data_start_row}"
            worksheet.sheet_view.showGridLines = False
            worksheet.page_setup.orientation = "landscape"
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = 0
            worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            workbook.active = workbook.index(worksheet)

    def _money_columns(self, headers):
        expected = {"versement", "debours", "honoraires", "tva", "honors_ht"}
        return [header for header in headers if self._normal_header(header) in expected]

    def _normal_header(self, header):
        text = str(header).strip().lower()
        replacements = {
            "°": "",
            "Â°": "",
            "é": "e",
            "Ã©": "e",
            "ç": "c",
            "Ã§": "c",
            ".": "",
            "/": "_",
            " ": "_",
            "'": "",
        }
        for old, new in replacements.items():
            text = text.replace(old, new)
        return text

    def _parse_money(self, value):
        text = str(value).strip()
        if not text or text.lower() == "nan":
            return None
        try:
            return float(text.replace(" ", "").replace(",", "."))
        except ValueError:
            return value

    def _export_column_width(self, header, values):
        preferred = {
            "nfact": 12,
            "rep": 14,
            "date": 14,
            "recu": 14,
            "libelles": 36,
            "versement": 16,
            "debours": 16,
            "honoraires": 16,
            "tva": 14,
            "honors_ht": 16,
        }
        normalized = self._normal_header(header)
        if normalized in preferred:
            return preferred[normalized]

        longest = len(str(header))
        for value in values:
            longest = max(longest, len(str(value)))
        return min(max(longest + 2, 12), 34)

    def _log_changes(self, old_df, new_df, headers):
        for row_index in range(min(len(old_df), len(new_df))):
            for column in headers:
                if column not in old_df.columns:
                    continue
                old_value = str(old_df.at[row_index, column]).strip()
                new_value = str(new_df.at[row_index, column]).strip()
                if old_value.endswith(".0"):
                    old_value = old_value[:-2]
                if new_value.endswith(".0"):
                    new_value = new_value[:-2]
                if old_value != new_value and not (old_value in ("nan", "") and new_value == ""):
                    database.log_edit(self.current_user, row_index + 2, column, old_value, new_value)

        for row_index in range(len(old_df), len(new_df)):
            database.log_edit(self.current_user, row_index + 2, "ROW", "", "ADDED")
